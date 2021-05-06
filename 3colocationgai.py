import openpyxl
path_1 = r'C:\Users\wzy\Desktop\data\data-1\data_colocation2_1.xlsx'
workbook = openpyxl.load_workbook(path_1)
sheetname = workbook.sheetnames
worksheet_3 = workbook['Sheet2']
worksheet_1 = workbook['Sheet1']
for i in range(1, worksheet_1.max_row):
    top_a = worksheet_1.cell(row=i, column=1)
    top_b = worksheet_1.cell(row=i, column=2)
    top_c = worksheet_1.cell(row=i, column=3)
    top_d = worksheet_1.cell(row=i, column=4)
    top_e = worksheet_1.cell(row=i, column=5)
    top_f = worksheet_1.cell(row=i, column=6)
    top_g = worksheet_1.cell(row=i, column=7)
    top_h = worksheet_1.cell(row=i, column=8)
    k = i+1
    bottom_a = worksheet_1.cell(row=k, column=1)
    bottom_b = worksheet_1.cell(row=k, column=2)
    bottom_c = worksheet_1.cell(row=k, column=3)
    bottom_d = worksheet_1.cell(row=k, column=4)
    bottom_e = worksheet_1.cell(row=k, column=5)
    bottom_f = worksheet_1.cell(row=k, column=6)
    bottom_g = worksheet_1.cell(row=k, column=7)
    bottom_h = worksheet_1.cell(row=k, column=8)
    while top_a.value == bottom_a.value and top_b.value == bottom_b.value and top_c.value == bottom_c.value and\
            top_d.value == bottom_d.value:
        if top_g.value != bottom_g.value:
            top = top_g.value
            bottom = bottom_g.value
            if top < bottom:
                sheetname_1 = '{a} and {b}'.format(a=top, b=bottom)
                worksheet_2 = workbook[sheetname_1]
                for m in range(1, worksheet_2.max_row):
                    data_1 = worksheet_2.cell(row=m, column=1)
                    data_2 = worksheet_2.cell(row=m, column=2)
                    data_3 = worksheet_2.cell(row=m, column=3)
                    data_4 = worksheet_2.cell(row=m, column=4)
                    data_5 = worksheet_2.cell(row=m, column=5)
                    data_6 = worksheet_2.cell(row=m, column=6)
                    data_7 = worksheet_2.cell(row=m, column=7)
                    data_8 = worksheet_2.cell(row=m, column=8)
                    # print(data_1.value, top_e.value, data_2.value, top_f.value, data_3.value, top_g.value, data_4.value
                    #       , top_h.value, data_5.value, bottom_e.value, data_6.value, bottom_f.value, data_7.value, bottom_g.value
                    #       , data_8.value, bottom_h.value)
                    if data_1.value == top_e.value and data_2.value == top_f.value and data_3.value == top_g.value and\
                            data_4.value == top_h.value and\
                            data_5.value == bottom_e.value and data_6.value == bottom_f.value and data_7.value == bottom_g.value and\
                            data_8.value == bottom_h.value:
                        lst = [top_a.value, top_b.value, top_c.value, top_d.value, data_1.value, data_2.value,
                               data_3.value, data_4.value, data_5.value, data_6.value, data_7.value, data_8.value]
                        worksheet_3.append(lst)
                        print(i, k, m)
            else:
                sheetname_1 = '{a} and {b}'.format(a=bottom, b=top)
                worksheet_2 = workbook[sheetname_1]
                for m in range(1, worksheet_2.max_row):
                    data_1 = worksheet_2.cell(row=m, column=1)
                    data_2 = worksheet_2.cell(row=m, column=2)
                    data_3 = worksheet_2.cell(row=m, column=3)
                    data_4 = worksheet_2.cell(row=m, column=4)
                    data_5 = worksheet_2.cell(row=m, column=5)
                    data_6 = worksheet_2.cell(row=m, column=6)
                    data_7 = worksheet_2.cell(row=m, column=7)
                    data_8 = worksheet_2.cell(row=m, column=8)
                    if data_1.value == bottom_e.value and data_2.value == bottom_f.value and data_3.value == bottom_g.value and \
                            data_4.value == bottom_h.value and\
                            data_5.value == top_e.value and data_6.value == top_f.value and data_7.value == top_g.value and \
                            data_8.value == top_h.value:
                        lst = [top_a.value, top_b.value, top_c.value, top_d.value, data_1.value, data_2.value,
                               data_3.value, data_4.value, data_5.value, data_6.value, data_7.value, data_8.value]
                        worksheet_3.append(lst)
                        print(i, k, m)
        k = k+1
        bottom_a = worksheet_1.cell(row=k, column=1)
        bottom_b = worksheet_1.cell(row=k, column=2)
        bottom_c = worksheet_1.cell(row=k, column=3)
        bottom_d = worksheet_1.cell(row=k, column=4)
        bottom_e = worksheet_1.cell(row=k, column=5)
        bottom_f = worksheet_1.cell(row=k, column=6)
        bottom_g = worksheet_1.cell(row=k, column=7)
        bottom_h = worksheet_1.cell(row=k, column=8)
workbook.save(path_1)






