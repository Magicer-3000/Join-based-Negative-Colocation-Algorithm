import openpyxl

path_1 = r'C:/Users/wzy/Desktop/data_2.xlsx'
path_2 = r'C:/Users/wzy/Desktop/data_2_1.xlsx'

workbook_2 = openpyxl.load_workbook(path_2)
worksheet_2 = workbook_2['Sheet1']

workbook_1 = openpyxl.load_workbook(path_1)                             # 打开表格
sheets = workbook_1.sheetnames
# 获取所有sheet表格

for i in range(1, 5):                                               # 循环1-8个表格
    worksheet_1 = workbook_1[sheets[i]]
    for m in range(1, worksheet_1.max_row-1):                               # m是当前点的行数
        top_a = worksheet_1.cell(row=m, column=1)                     # 获取一个点的X,Y
        top_b = worksheet_1.cell(row=m, column=2)
        top_c = worksheet_1.cell(row=m, column=3)
        top_d = worksheet_1.cell(row=m, column=4)
        for k in range(i+1, 5):                                     # 进去另一个空间特征获取点
            worksheet_3 = workbook_1[sheets[k]]
            for j in range(1, worksheet_3.max_row-1):                      # 获取将要进行对比的点的X,Y
                bottom_a = worksheet_3.cell(row=j, column=1)
                bottom_b = worksheet_3.cell(row=j, column=2)
                bottom_c = worksheet_3.cell(row=j, column=3)
                bottom_d = worksheet_3.cell(row=j, column=4)
                if (float(top_a.value)-float(bottom_a.value)) ** 2 + (float(top_b.value)-float(bottom_b.value)) ** 2 < 1.0:
                    lst = [top_a.value, top_b.value, top_c.value, top_d.value,
                           bottom_a.value, bottom_b.value, bottom_c.value, bottom_d.value]
                    worksheet_2.append(lst)
                    print(i, m, k, j)
workbook_2.save(path_2)
