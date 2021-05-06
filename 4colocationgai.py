import openpyxl
path_1 = r'C:\Users\Admin\Desktop\data\data-1\data_colocation3_1.xlsx'
workbook = openpyxl.load_workbook(path_1)
sheetname = workbook.sheetnames
worksheet_3 = workbook['Sheet2']
worksheet_1 = workbook['Sheet1']
top = bottom = 0
for i in range(1, worksheet_1.max_row):
    top_a = worksheet_1.cell(row=i, column=1)
    top_b = worksheet_1.cell(row=i, column=2)
    top_c = worksheet_1.cell(row=i, column=3)
    top_d = worksheet_1.cell(row=i, column=4)
    top_e = worksheet_1.cell(row=i, column=5)
    top_f = worksheet_1.cell(row=i, column=6)
    top_g = worksheet_1.cell(row=i, column=7)
    top_h = worksheet_1.cell(row=i, column=8)
    top_i = worksheet_1.cell(row=i, column=9)
    top_j = worksheet_1.cell(row=i, column=10)
    top_k = worksheet_1.cell(row=i, column=11)
    top_l = worksheet_1.cell(row=i, column=12)
    k = i+1
    bottom_a = worksheet_1.cell(row=k, column=1)
    bottom_b = worksheet_1.cell(row=k, column=2)
    bottom_c = worksheet_1.cell(row=k, column=3)
    bottom_d = worksheet_1.cell(row=k, column=4)
    bottom_e = worksheet_1.cell(row=k, column=5)
    bottom_f = worksheet_1.cell(row=k, column=6)
    bottom_g = worksheet_1.cell(row=k, column=7)
    bottom_h = worksheet_1.cell(row=k, column=8)
    bottom_i = worksheet_1.cell(row=k, column=9)
    bottom_j = worksheet_1.cell(row=k, column=10)
    bottom_k = worksheet_1.cell(row=k, column=11)
    bottom_l = worksheet_1.cell(row=k, column=12)
    while top_a.value == bottom_a.value and top_b.value == bottom_b.value and top_c.value == bottom_c.value and\
            top_d.value == bottom_d.value and top_e.value == bottom_e.value and top_f.value == bottom_f.value and\
            top_g.value == bottom_g.value and top_h.value == bottom_h.value:
        if top_k.value != bottom_k.value:
            if top_k.value < bottom_k.value:
                top = top_k.value
                bottom = bottom_k.value
            elif top_k.value > bottom_k.value:
                bottom = top_k.value
                top = bottom_k.value
            sheetname_1 = '{a} and {b} and {c}'.format(a=top_g.value, b=top, c=bottom)
            worksheet_2 = workbook[sheetname_1]
            for m in range(1, worksheet_2.max_row):
                data_1 = worksheet_2.cell(row=m, column=1)
                data_2 = worksheet_2.cell(row=m, column=2)
                data_3 = worksheet_2.cell(row=m, column=3)
                data_4 = worksheet_2.cell(row=m, column=4)
                data_5 = worksheet_2.cell(row=m, column=5)
                data_6 = worksheet_2.cell(row=m, column=6)
                data_7 = worksheet_2.cell(row=m, column=7)
                data_8 = worksheet_2.cell(row=m, column=8)
                data_9 = worksheet_2.cell(row=m, column=9)
                data_10 = worksheet_2.cell(row=m, column=10)
                data_11 = worksheet_2.cell(row=m, column=11)
                data_12 = worksheet_2.cell(row=m, column=12)
                if data_1.value == top_e.value and data_2.value == top_f.value and data_3.value == top_g.value and\
                        data_4.value == top_h.value and\
                        data_5.value == top_i.value and data_6.value == top_j.value and\
                        data_7.value == top_k.value and data_8.value == top_l.value and \
                        data_9.value == bottom_i.value and data_10.value == bottom_j.value and\
                        data_11.value == bottom_k.value and data_12.value == bottom_l.value:
                    lst = [top_a.value, top_b.value, top_c.value, top_d.value, data_1.value, data_2.value,
                           data_3.value, data_4.value, data_5.value, data_6.value, data_7.value, data_8.value,
                           data_9.value, data_10.value, data_11.value, data_12.value]
                    worksheet_3.append(lst)
                    print(i, k, m)
            # else:
            #     sheetname_1 = '{a} and {b} and {c}'.format(a=top_g.value, b=bottom, c=top)
            #     worksheet_2 = workbook[sheetname_1]
            #     for m in range(1, worksheet_2.max_row):
            #         data_1 = worksheet_2.cell(row=m, column=1)
            #         data_2 = worksheet_2.cell(row=m, column=2)
            #         data_3 = worksheet_2.cell(row=m, column=3)
            #         data_4 = worksheet_2.cell(row=m, column=4)
            #         data_5 = worksheet_2.cell(row=m, column=5)
            #         data_6 = worksheet_2.cell(row=m, column=6)
            #         data_7 = worksheet_2.cell(row=m, column=7)
            #         data_8 = worksheet_2.cell(row=m, column=8)
            #         data_9 = worksheet_2.cell(row=m, column=9)
            #         data_10 = worksheet_2.cell(row=m, column=10)
            #         data_11 = worksheet_2.cell(row=m, column=11)
            #         data_12 = worksheet_2.cell(row=m, column=12)
            #         if data_1.value == top_e.value and data_2.value == top_f.value and data_3.value == top_g.value and \
            #                 data_4.value == top_h.value and \
            #                 data_5.value == bottom_i.value and data_6.value == bottom_j.value and \
            #                 data_7.value == bottom_k.value and data_8.value == bottom_l.value and \
            #                 data_9.value == top_i.value and data_10.value == top_j.value and \
            #                 data_11.value == top_k.value and data_8.value == top_l.value:
            #             lst = [top_a.value, top_b.value, top_c.value, top_d.value, data_1.value, data_2.value,
            #                    data_3.value, data_4.value, data_5.value, data_6.value, data_7.value, data_8.value,
            #                    data_9.value, data_10.value, data_11.value, data_12.value]
            #             worksheet_3.append(lst)
            #             print(i, k, m)
        k = k+1
        bottom_a = worksheet_1.cell(row=k, column=1)
        bottom_b = worksheet_1.cell(row=k, column=2)
        bottom_c = worksheet_1.cell(row=k, column=3)
        bottom_d = worksheet_1.cell(row=k, column=4)
        bottom_e = worksheet_1.cell(row=k, column=5)
        bottom_f = worksheet_1.cell(row=k, column=6)
        bottom_g = worksheet_1.cell(row=k, column=7)
        bottom_h = worksheet_1.cell(row=k, column=8)
        bottom_i = worksheet_1.cell(row=k, column=9)
        bottom_j = worksheet_1.cell(row=k, column=10)
        bottom_k = worksheet_1.cell(row=k, column=11)
        bottom_l = worksheet_1.cell(row=k, column=12)
workbook.save(path_1)






