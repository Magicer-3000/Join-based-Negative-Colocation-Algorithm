import openpyxl
path = r''
workbook = openpyxl.load_workbook(path)
sheetsname = workbook.sheetnames
prevalent_size_3 = []
prevalent_size_4 = []
min_prev = 0.6

worksheet_positive_size_2 = workbook['positive_size_2']
worksheet_negative_size_2 = workbook['nagative_size_2']
for i in range(1, worksheet_positive_size_2.max_row):
    positive_feature_1 = worksheet_positive_size_2.cell(row=i, cloumn=1)
    positive_feature_2 = worksheet_positive_size_2.cell(row=i, cloumn=2)
    for k in range(1, worksheet_negative_size_2.max_row):
        nagative_feature_1 = worksheet_negative_size_2.cell(row=k, cloumn=1)
        nagative_feature_2 = worksheet_negative_size_2.cell(row=k, cloumn=2)
        if nagative_feature_1 == positive_feature_1 or nagative_feature_1 == positive_feature_2 or nagative_feature_2 == positive_feature_1 or\
           nagative_feature_2 == positive_feature_2:
            nagative_size_3 = combine(positive_feature_1, positive_feature_2, nagative_feature_1, nagative_feature_2)
            nagative_size_3_PI = calculate_PI(nagative_size_3)
            if nagative_size_3_PI >= min_prev:
                prevalent_size_3.append(nagative_size_3)






